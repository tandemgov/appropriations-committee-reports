"""Per-report Excel workbooks a staffer can audit without trusting us.

The premise: an appropriations staffer will not adopt a parsed dataset because a pipeline
printed "verified". They will adopt it when they can perform, in thirty seconds and with the
tool already on their desk, the check they would otherwise do against the PDF -- highlight the
account rows, read the sum off the status bar, compare it to the recap.

So the workbook ships the arithmetic, not a claim about the arithmetic:

* ``computed`` on every total row is a live ``=SUM()`` over the exact cells that total consumed.
  Nothing is precomputed. Delete a line item and the total goes red; Excel, not this program,
  is doing the adding.
* ``check`` is ``printed - computed``, conditionally formatted. Zero is the assertion.
* Non-add memo rows are greyed and their amounts sit in a column no ``SUM`` touches, so the
  reason they are excluded is legible rather than hidden in a filter.
* Leaves and rollups live in separate columns, which is what lets a parent total sum its child
  subtotals without double-counting the leaves underneath them.

A staffer who does not believe the ``check`` column can select the leaf cells by hand and read
Excel's own status bar. That is the point: every number in here is either lifted verbatim from
the report or derived by a formula they can see.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from approps.verification.reconcile import (
    PRIMARY_COLUMN,
    ReportReconciliation,
    Status,
    is_memo,
    is_total,
    recover_primary,
)

MONEY = "#,##0;[Red]-#,##0;—"

_HEADERS = [
    "line item",
    "amount",
    "leaf",
    "rollup",
    "computed",
    "check",
    "status",
    "note",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
_MEMO_FONT = Font(color="9C9C9C", italic=True)
_OK_FILL = PatternFill("solid", fgColor="C6EFCE")
_BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
_THIN = Side(style="thin", color="BFBFBF")


def _compress(cells: list[int]) -> str:
    """Contiguous row numbers collapse into ranges, so a 60-child total reads as C5:C64."""
    if not cells:
        return ""
    runs: list[tuple[int, int]] = []
    start = previous = cells[0]
    for row in cells[1:]:
        if row == previous + 1:
            previous = row
            continue
        runs.append((start, previous))
        start = previous = row
    runs.append((start, previous))
    return ",".join(f"{a}" if a == b else f"{a}:{b}" for a, b in runs)


def _sum_formula(leaf_rows: list[int], rollup_rows: list[int]) -> str:
    """A SUM over the child cells, in whichever column each child lives.

    Children are addressed individually rather than as one sweeping range because a parent
    total's children are its child *subtotals* plus any loose leaves -- and the leaves those
    subtotals already absorbed sit between them. Summing the range would count them twice.
    """
    parts = []
    if leaf_rows:
        parts += [f"C{ref}" if ":" not in ref else f"C{ref.split(':')[0]}:C{ref.split(':')[1]}"
                  for ref in _compress(leaf_rows).split(",")]
    if rollup_rows:
        parts += [f"D{ref}" if ":" not in ref else f"D{ref.split(':')[0]}:D{ref.split(':')[1]}"
                  for ref in _compress(rollup_rows).split(",")]
    return f"=SUM({','.join(parts)})" if parts else "=0"


def _write_line_items(
    sheet, rows: list[dict], result: ReportReconciliation, sheet_row_of: dict[int, int]
) -> None:
    sheet.append(_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    checks_by_index = {c.index: c for c in result.checks}
    # Exactly the rows some total's SUM will reference. A memo the printed total endorsed is in
    # here; one it excluded is not, and so gets no summable cell at all -- its exclusion is
    # visible on the page rather than hidden in a filter. Keyed on child membership rather than
    # on memo_mode, because a row can be *both* a memo and a subtotal (a fee-funded account
    # printed in parentheses), in which case it is a mandatory rollup child regardless.
    children_of_any = {i for c in result.checks for i in c.child_indices}

    for index, row in enumerate(rows):
        sheet_row = index + 2
        sheet_row_of[index] = sheet_row
        label = row.get("line_item_text") or ""
        # The amount the arithmetic uses, which is what the reader must see. Where the report
        # printed a dot leader, `recover_primary` reconstructs the level from the row's own
        # delta columns; showing the raw blank instead would make a reconciling total look red.
        amount = recover_primary(row)
        total = is_total(row)
        memo = is_memo(row)
        summable = not memo or index in children_of_any

        sheet.cell(sheet_row, 1, label)
        sheet.cell(sheet_row, 2, amount)
        if row.get(PRIMARY_COLUMN) is None and amount is not None:
            sheet.cell(sheet_row, 8, "recovered from this row's deltas")

        # The column is chosen by what the row *is* -- leaf or rollup -- which is exactly how
        # _sum_formula addresses it. Memo-ness decides only whether it gets a column at all.
        if summable:
            sheet.cell(sheet_row, 4 if total else 3, amount)

        if memo:
            for column in range(1, 9):
                sheet.cell(sheet_row, column).font = _MEMO_FONT
            sheet.cell(
                sheet_row,
                8,
                "memo — the printed total adds it"
                if index in children_of_any
                else "non-add memo — no SUM reaches it",
            )

        for column in (2, 3, 4, 5, 6):
            sheet.cell(sheet_row, column).number_format = MONEY

    for check in result.checks:
        sheet_row = sheet_row_of[check.index]
        for column in range(1, 9):
            cell = sheet.cell(sheet_row, column)
            cell.fill = _TOTAL_FILL
            cell.border = Border(top=_THIN)
            if column == 1:
                cell.font = Font(bold=True)

        sheet.cell(sheet_row, 7, check.status.value)

        if check.status is Status.UNCHECKED or not check.child_indices:
            continue

        leaf_rows = sorted(
            sheet_row_of[i] for i in check.child_indices if i not in checks_by_index
        )
        rollup_rows = sorted(sheet_row_of[i] for i in check.child_indices if i in checks_by_index)

        computed = sheet.cell(sheet_row, 5, _sum_formula(leaf_rows, rollup_rows))
        computed.number_format = MONEY
        check_cell = sheet.cell(sheet_row, 6, f"=B{sheet_row}-E{sheet_row}")
        check_cell.number_format = MONEY

    last = len(rows) + 1
    sheet.conditional_formatting.add(
        f"F2:F{last}",
        CellIsRule(operator="equal", formula=["0"], fill=_OK_FILL),
    )
    sheet.conditional_formatting.add(
        f"F2:F{last}",
        CellIsRule(operator="notEqual", formula=["0"], fill=_BAD_FILL),
    )

    sheet.freeze_panes = "B2"
    sheet.column_dimensions["A"].width = 62
    for letter in "BCDEF":
        sheet.column_dimensions[letter].width = 16
    sheet.column_dimensions["G"].width = 18
    sheet.column_dimensions["H"].width = 34
    sheet.auto_filter.ref = f"A1:H{last}"


def _write_reconciliation(sheet, result: ReportReconciliation, sheet_row_of: dict[int, int]) -> None:
    headers = ["printed total", "row", "printed", "computed", "check", "children", "status"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL

    for offset, check in enumerate(result.checks, start=2):
        source_row = sheet_row_of[check.index]
        sheet.cell(offset, 1, check.label)
        link = sheet.cell(offset, 2, f"='Line items'!A{source_row}")
        link.alignment = Alignment(horizontal="left")
        sheet.cell(offset, 3, f"='Line items'!B{source_row}").number_format = MONEY
        if check.status is not Status.UNCHECKED and check.child_indices:
            sheet.cell(offset, 4, f"='Line items'!E{source_row}").number_format = MONEY
            sheet.cell(offset, 5, f"='Line items'!F{source_row}").number_format = MONEY
        sheet.cell(offset, 6, len(check.child_indices))
        sheet.cell(offset, 7, check.status.value)

    last = len(result.checks) + 1
    sheet.conditional_formatting.add(
        f"E2:E{last}", CellIsRule(operator="equal", formula=["0"], fill=_OK_FILL)
    )
    sheet.conditional_formatting.add(
        f"E2:E{last}", CellIsRule(operator="notEqual", formula=["0"], fill=_BAD_FILL)
    )
    sheet.column_dimensions["A"].width = 54
    sheet.column_dimensions["B"].width = 40
    for letter in "CDE":
        sheet.column_dimensions[letter].width = 16
    sheet.column_dimensions["G"].width = 18
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{last}"


def _write_readme(sheet, result: ReportReconciliation) -> None:
    strict = result.strict_pass_rate
    lines = [
        (f"{result.report_id} — reconciliation workbook", True),
        ("", False),
        ("Every 'computed' cell is a live =SUM() over the line items above it.", False),
        ("Nothing on this sheet was precomputed: Excel does the adding, not the extractor.", False),
        ("", False),
        ("How to check a number the way you would against the PDF:", True),
        ("  1. Find the total you care about on the 'Reconciliation' sheet.", False),
        ("  2. Click through to it on 'Line items'.", False),
        ("  3. Select the 'leaf' cells above it. Excel's status bar shows their sum.", False),
        ("  4. It equals the 'printed' cell, which is lifted verbatim from the report.", False),
        ("", False),
        ("Columns", True),
        ("  amount    as the committee printed it. Where the report printed a dot leader,", False),
        ("            this is reconstructed from the row's own delta columns and the", False),
        ("            'note' column says so.", False),
        ("  leaf      a line item's amount; the only column a subtotal's SUM reaches", False),
        ("  rollup    a subtotal's amount, so a parent can sum subtotals without", False),
        ("            double-counting the leaves they already absorbed", False),
        ("  computed  =SUM() of this total's children", False),
        ("  check     amount - computed. Zero is the assertion. Green is good.", False),
        ("", False),
        ("Greyed italic rows are parenthesized memos — limitations, transfer authority,", False),
        ("'of which' breakouts. A parenthesis never means a negative here; real negatives", False),
        ("print a minus sign. Whether a memo is summed is decided by the printed total,", False),
        ("not by us:", False),
        ("  'non-add memo'          already counted inside a sibling line. No SUM reaches", False),
        ("                          it — its amount sits in no summable column.", False),
        ("  'memo — added to total' the printed total only closes with it added, so it is", False),
        ("                          real additional budget authority at this level.", False),
        ("", False),
        ("Totals that cannot be checked by summing children:", True),
        ("  overlapping_view  advance-appropriation / forward-funding totals, which", False),
        ("                    re-aggregate rows already counted under another view", False),
        ("  unchecked         the report printed a dot leader, not an amount", False),
        ("", False),
        (f"totals: {len(result.checks)}   tie exactly: {result.n_ok}   "
         f"genuine failures: {result.n_genuine_failures}", True),
        (f"strict pass rate (excludes overlapping views): "
         f"{strict:.1%}" if strict is not None else "strict pass rate: n/a", True),
    ]
    for offset, (text, bold) in enumerate(lines, start=1):
        cell = sheet.cell(offset, 1, text)
        if bold:
            cell.font = Font(bold=True)
    sheet.column_dimensions["A"].width = 96


def write_report_workbook(path: Path, rows: list[dict], result: ReportReconciliation) -> Path:
    """Write one report's reconciliation workbook. Returns the path written."""
    workbook = Workbook()
    readme = workbook.active
    readme.title = "Read me"
    _write_readme(readme, result)

    sheet_row_of: dict[int, int] = {}
    _write_line_items(workbook.create_sheet("Line items"), rows, result, sheet_row_of)
    _write_reconciliation(workbook.create_sheet("Reconciliation"), result, sheet_row_of)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path
